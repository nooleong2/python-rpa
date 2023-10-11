from openpyxl import load_workbook
# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

# # 수식 그대로 출력
# for row in ws.values:
#     for cell in row:
#         print(cell)

wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

# 수식이 아닌 실제 데이터를 가지고 옴
# evaluate 되지 않은 상태의 데이터는 None 이라고 표시
# excel을 열어서 다시 한번 저장하고 난 후에는 계산된 결과 값을 받을 수 있음
for row in ws.values:
    for cell in row:
        print(cell)