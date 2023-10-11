from openpyxl import Workbook
from random import *
from openpyxl.utils.cell import coordinate_from_string

wb = Workbook()
ws = wb.active

# 한 줄씩 데이터 넣기
ws.append(["number", "eng", "math"])
for i in range(1, 11): # 10개 데이터 넣기
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"] # 영어 column만 가져 오기
print(col_B) # ws["B"]의 정보

for cell in col_B:
    print(cell.value)

col_range = ws["B:C"] # B, C column 함께 가져오기
for cols in col_range: # B Column 부터 순차적으로
    for cell in cols: # 컬럼의 셀을 하나씩 돌고
        print(cell.value) # 값을 하나씩 출력

row_title = ws[1] # 1번째 row만 가지고 오기
for cell in row_title:
    print(cell.value)

row_range = ws[2:6] # 2 ~ 6번째 줄까지 가져오기
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()

row_range2 = ws[2:ws.max_row] # 2 ~ 마지막 줄까지 가져오기
for rows in row_range2:
    for cell in rows:
        print(cell.value, end=" ") # 각 cell의 값 가져오기
        print(cell.coordinate, end=" ") # 각 cell의 좌표 정보 가져옴 import coordinate_from_string 해야 사용 가능
        xy = coordinate_from_string(cell.coordinate) # A/1, AZ/255 형태로 만들어 줌
        print(xy, end=" ") # 튜플 형태로 ('A', 2), ('B', 2)
        print(xy[0], end="") # A
        print(xy[1], end=" ") # 1
    print()

# 전체 rows
print(ws.rows) # ws.rows의 정보
print(tuple(ws.rows)) # tuple 형태의 정보
for row in tuple(ws.rows):
    print(row)
    print(row[0].value)

# 전체 columns
print(ws.columns)
print(tuple(ws.columns))
for col in tuple(ws.columns):
    print(col)
    print(col[0].value)

# 전체 row 가죠오기
for row in ws.iter_rows():
    print(row[1].value)

# 전체 column 가져오기
for col in ws.iter_cols():
    print(col[0].value)

# 범위 지정 cell 가져오기
# 1 ~ 5번째 줄까지, 2 ~ 3 열까지
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
    print(row[0].value, row[1].value)

for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
    print(col)

wb.save("sample.xlsx")
wb.close()