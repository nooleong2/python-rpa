from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# 지정한 범위 위치 이동
# number eng math
# number (kor) eng math

# ws.move_range("B1:C11", rows=0, cols=1) # 이동할 데이터 범위, rows 이동 칸, cols 이동 칸
# ws["B1"].value = "kor" # B1 cell에 kor 입력

ws.move_range("C1:C11", rows=5, cols=-1)

wb.save("sample_korea_add.xlsx")
wb.close()