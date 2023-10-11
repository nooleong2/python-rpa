from openpyxl import load_workbook

wb = load_workbook("sample.xlsx") # sample.xlsx 파일을 열기
ws = wb.active # 활성화된 시트 사용

for row in ws.iter_rows(min_row=2):
    if int(row[1].value) > 80:
        print(row[0].value, "번 학생은 영어 천재")

for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "eng":
            cell.value = "computer"

wb.save("sample_modified.xlsx")
wb.close()