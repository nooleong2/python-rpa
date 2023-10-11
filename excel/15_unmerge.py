from openpyxl import load_workbook
wb = load_workbook("sample_merge.xlsx")
ws = wb.active

# B2 ~ D2 까지 병합 되어 있던 셀 해제
ws.unmerge_cells("B2:D2")

wb.save("sample_unmerge.xlsx")
wb.close()