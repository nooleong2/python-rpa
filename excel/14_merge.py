from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# 병합
ws.merge_cells("B2:D2") # B2 ~ D2 까지 합침
ws["B2"].value = "Merged Cell"

wb.save("sample_merge.xlsx")
wb.close()