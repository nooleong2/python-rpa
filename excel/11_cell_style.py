from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")
ws = wb.active

# number, eng, math
a1 = ws["A1"] # number
b1 = ws["B1"] # eng
c1 = ws["C1"] # math

# A 열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5

# 1 행의 높이를 50으로 설정
ws.row_dimensions[1].height = 50

# 스타일 적용
a1.font = Font(color="ff0000", italic=True, bold=True) # 글자 색상, 이텔릭체, 볼드
b1.font = Font(color="cc33ff", name="Arial", strike=True) # 글자 색상, 폰트명 지정, 글 취소 선
c1.font = Font(color="0000ff", size=20, underline="single") # 글자 크기 20, 밑줄 적용

# 외각선(테두리) 작용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 80점 넘는 셀에 대해서 초록색으로 적용
for row in ws.rows:
    for cell in row:
        #각 cell에 대해서 정렬 (center, left, right, top, bottom)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if cell.column == 1: # A 번호열은 제외
            continue

        # cell 이 정수형 데이터이고 80점 보다 높으면
        if isinstance(cell.value, int) and cell.value >= 80:
            cell.fill = PatternFill(fgColor="00ff00", fill_type="solid") # 배경색 초록
            cell.font = Font(color="ff0000") # 폰트 색상 변경

# 틀 고정 (ex : 제목이되는 row 부분)
ws.freeze_panes = "B2" # B2 기준으로 틀 고정

wb.save("sample_style.xlsx")
wb.close()