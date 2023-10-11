from openpyxl import Workbook

wb = Workbook() # 워크북 생성
ws = wb.create_sheet() # 새로운 시트 기본이름으로 생성
ws.title = "MySheet" # 시트 이름변경
ws.sheet_properties.tabColor = "ff66ff" # rgb 형태로 값을 넣어주면 탭 색상 변경

# Sheet, MySheet, YourSheet
ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 시트 생성
ws2 = wb.create_sheet("NewSheet", 2) # 시트생성 위치 지정

new_ws = wb["NewSheet"] # 딕셔너리 형태로 Sheet에 접근

print(wb.sheetnames) # 모든 sheet 이름

# 시트 복사
new_ws["A1"] = "Test" # A1 컬럼에 텍스트 삽입
target = wb.copy_worksheet(new_ws) # new_ws 시트 복사 후 맨 뒤에 추가
target.title = "Copied Sheet" # 카피 된 시트의 시트명 변경

wb.save("sample.xlsx")
wb.close()
